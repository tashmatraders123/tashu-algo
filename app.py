"""
app.py
Multi-user web dashboard backend. Each person who signs in gets their own
isolated bot process, their own Delta API keys (never visible to anyone
else), and their own settings.

IMPORTANT: this app has no email verification and is meant for a small
group of people you personally trust with the link to this site --
anyone who can reach it can create an account. Do not publicize the URL
widely without adding stronger access control first.
"""
import io
import json
import os
import secrets
import subprocess
import sys
import threading
import time
from functools import wraps

from flask import Flask, jsonify, request, session, redirect, url_for, render_template, send_file

import config
import support
import users
from delta_api import DeltaClient, DeltaAPIError

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BOT_SCRIPT = os.path.join(BASE_DIR, "bot.py")
SECRET_KEY_PATH = os.path.join(BASE_DIR, "secret.key")

COMMON_SYMBOLS = [
    "BTCUSD", "ETHUSD", "SOLUSD", "XRPUSD", "DOGEUSD",
    "ADAUSD", "POLUSD", "LTCUSD", "BNBUSD", "AVAXUSD",
]

app = Flask(
    __name__,
    static_folder=os.path.join(BASE_DIR, "web"),
    static_url_path="/static",
    template_folder=os.path.join(BASE_DIR, "templates"),
)

# Persistent session secret (generated once, reused across restarts so
# people don't get logged out every time the server restarts)
if os.path.exists(SECRET_KEY_PATH):
    with open(SECRET_KEY_PATH, "r") as f:
        app.secret_key = f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    with open(SECRET_KEY_PATH, "w") as f:
        f.write(app.secret_key)

# username -> {"process": Popen, "stop_requested_at": float|None}
_bot_processes = {}
_processes_lock = threading.Lock()


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Not logged in"}), 401
            return redirect(url_for("login_page"))
        if users.is_suspended(session["username"]):
            # Access was revoked after this session started -- log them
            # out now rather than letting an already-open tab keep working.
            session.pop("username", None)
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Your access has been suspended"}), 403
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login_page"))
        if not users.is_admin(session["username"]):
            return "Admins only.", 403
        return f(*args, **kwargs)
    return wrapper


def user_paths(username):
    d = users.user_dir(username)
    return {
        "log": os.path.join(d, "bot.log"),
        "stop_flag": os.path.join(d, "stop.flag"),
        "log_rel": os.path.relpath(os.path.join(d, "bot.log"), BASE_DIR),
        "stop_flag_rel": os.path.relpath(os.path.join(d, "stop.flag"), BASE_DIR),
    }


# ----------------------------------------------------------------------
# Auth pages
# ----------------------------------------------------------------------
@app.route("/register", methods=["GET", "POST"])
def register_page():
    if request.method == "GET":
        return render_template("register.html", error=None)
    username = request.form.get("username", "")
    email = request.form.get("email", "")
    password = request.form.get("password", "")
    confirm_password = request.form.get("confirm_password", "")

    if password != confirm_password:
        return render_template("register.html", error="Passwords do not match")

    ok, error = users.create_user(username, password, email=email)
    if not ok:
        return render_template("register.html", error=error)
    if users.is_admin(username.strip().lower()):
        # first-ever user, auto-approved as admin
        session["username"] = username.strip().lower()
        return redirect(url_for("dashboard"))
    return render_template(
        "login.html",
        error=None,
        info="Account created! An admin needs to approve it before you can log in.",
    )


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "GET":
        return render_template("login.html", error=None)
    username = request.form.get("username", "")
    password = request.form.get("password", "")
    ok, reason = users.verify_login(username, password)
    if ok:
        session["username"] = username.strip().lower()
        return redirect(url_for("dashboard"))
    if reason == "pending":
        return render_template(
            "login.html", error="Your account is waiting for admin approval. Check back soon."
        )
    if reason == "suspended":
        return render_template(
            "login.html", error="Your access has been suspended. Contact the admin for details."
        )
    return render_template("login.html", error="Incorrect username or password")


@app.route("/logout")
def logout():
    session.pop("username", None)
    return redirect(url_for("login_page"))


@app.route("/")
@login_required
def dashboard():
    return render_template(
        "index.html",
        username=session["username"],
        common_symbols=COMMON_SYMBOLS,
        is_admin=users.is_admin(session["username"]),
    )


@app.route("/admin")
@admin_required
def admin_page():
    return render_template(
        "admin.html",
        username=session["username"],
        pending=users.list_pending_users(),
        all_users=users.list_all_users(),
    )


@app.route("/admin/approve", methods=["POST"])
@admin_required
def admin_approve():
    target = request.form.get("username", "")
    users.approve_user(target)
    return redirect(url_for("admin_page"))


@app.route("/admin/reject", methods=["POST"])
@admin_required
def admin_reject():
    target = request.form.get("username", "")
    users.reject_user(target)
    return redirect(url_for("admin_page"))


@app.route("/admin/suspend", methods=["POST"])
@admin_required
def admin_suspend():
    target = request.form.get("username", "").strip().lower()
    users.suspend_user(target)
    # If their bot is currently running, stop it immediately rather than
    # waiting for them to notice they've lost access.
    with _processes_lock:
        entry = _bot_processes.get(target)
        if entry is not None and entry["process"].poll() is None:
            entry["stop_requested_at"] = time.time()
    paths = user_paths(target)
    try:
        with open(paths["stop_flag"], "w") as f:
            f.write("stop")
    except OSError:
        pass
    return redirect(url_for("admin_page"))


@app.route("/admin/unsuspend", methods=["POST"])
@admin_required
def admin_unsuspend():
    target = request.form.get("username", "")
    users.unsuspend_user(target)
    return redirect(url_for("admin_page"))


# ----------------------------------------------------------------------
# API: status
# ----------------------------------------------------------------------
@app.route("/api/status")
@login_required
def api_status():
    username = session["username"]
    with _processes_lock:
        entry = _bot_processes.get(username)
        running = entry is not None and entry["process"].poll() is None
        if entry is not None and not running:
            del _bot_processes[username]

    settings = users.get_settings(username)
    return jsonify({
        "running": running,
        "username": username,
        "symbol": settings["symbol"],
        "fixed_size": settings["fixed_size"],
        "dry_run": settings["dry_run"],
        "use_testnet": settings["use_testnet"],
        "api_key_masked": users.mask_secret(settings["api_key"]),
        "has_api_keys": bool(settings["api_key"] and settings["api_secret"]),
        "common_symbols": COMMON_SYMBOLS,
    })


@app.route("/api/account")
@login_required
def api_account():
    username = session["username"]
    settings = users.get_settings(username)
    if not settings["api_key"] or not settings["api_secret"]:
        return jsonify({"ok": False, "error": "Add your Delta API key/secret in Settings first"})
    try:
        base_url = (
            "https://cdn-ind.testnet.deltaex.org" if settings["use_testnet"]
            else "https://api.india.delta.exchange"
        )
        client = DeltaClient(base_url=base_url, api_key=settings["api_key"], api_secret=settings["api_secret"])
        balance = client.get_available_balance()
        product = client.get_product(settings["symbol"])
        positions = client.get_positions(product["id"])
        if isinstance(positions, dict):
            positions = [positions]
        open_position = None
        for p in positions:
            if float(p.get("size", 0) or 0) != 0:
                open_position = {
                    "size": p.get("size"),
                    "entry_price": p.get("entry_price"),
                    "unrealized_pnl": p.get("unrealized_pnl"),
                }
                break
        return jsonify({"ok": True, "balance": balance, "position": open_position})
    except DeltaAPIError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Unexpected error: {e}"})


def _read_position_state(username):
    """Reads the bot's own tracked trade state (SL/trailing stop, stage,
    R-tracking data) written by the running bot subprocess. Returns None
    if the bot isn't tracking a trade or hasn't run yet."""
    path = os.path.join(users.user_dir(username), "position_state.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


@app.route("/api/position")
@login_required
def api_position():
    """
    Rich position-details view for the dashboard: merges the LIVE position
    straight from Delta (real entry price, size, unrealized P&L) with the
    bot's own tracked state (current stop / trailing stage / R-distance),
    since the exchange itself has no concept of "which stage of our
    strategy this trade is in."
    """
    username = session["username"]
    settings = users.get_settings(username)
    state = _read_position_state(username)

    result = {"ok": True, "has_position": False, "symbol": settings["symbol"]}

    if not settings["api_key"] or not settings["api_secret"]:
        result["ok"] = False
        result["error"] = "Add your Delta API key/secret first"
        return jsonify(result)

    try:
        base_url = (
            "https://cdn-ind.testnet.deltaex.org" if settings["use_testnet"]
            else "https://api.india.delta.exchange"
        )
        client = DeltaClient(base_url=base_url, api_key=settings["api_key"], api_secret=settings["api_secret"])
        result["balance"] = client.get_available_balance()

        product = client.get_product(settings["symbol"])
        positions = client.get_positions(product["id"])
        if isinstance(positions, dict):
            positions = [positions]
        live_position = next((p for p in positions if float(p.get("size", 0) or 0) != 0), None)

        try:
            result["current_price"] = float(client.get_ticker(settings["symbol"])["close"])
        except DeltaAPIError:
            result["current_price"] = None

        if live_position:
            result["has_position"] = True
            result["side"] = "long" if float(live_position.get("size", 0)) > 0 else "short"
            result["size"] = live_position.get("size")
            result["entry_price"] = live_position.get("entry_price")
            result["unrealized_pnl"] = live_position.get("unrealized_pnl")
        elif state:
            # Dry-run / just-filled timing gap: bot is tracking a trade the
            # exchange hasn't reflected yet (or never will, in dry-run).
            result["has_position"] = True
            result["side"] = state.get("direction")
            result["size"] = state.get("size")
            result["entry_price"] = state.get("entry_price")
            result["unrealized_pnl"] = None

        if state:
            result["stage"] = state.get("stage")
            result["current_stop"] = state.get("current_stop")
            result["sl_distance"] = state.get("sl_distance")
            result["opened_at"] = state.get("opened_at")

            sl_dist = state.get("sl_distance")
            entry = result.get("entry_price")
            cur = result.get("current_price")
            direction = result.get("side")
            if sl_dist and entry is not None and cur is not None and direction:
                try:
                    entry_f, cur_f = float(entry), float(cur)
                    if direction == "long":
                        result["r_multiple"] = (cur_f - entry_f) / sl_dist
                    else:
                        result["r_multiple"] = (entry_f - cur_f) / sl_dist
                except (TypeError, ZeroDivisionError):
                    pass

        return jsonify(result)
    except DeltaAPIError as e:
        result["ok"] = False
        result["error"] = str(e)
        return jsonify(result)
    except Exception as e:
        result["ok"] = False
        result["error"] = f"Unexpected error: {e}"
        return jsonify(result)


_market_data_client = None


def _get_market_data_client():
    global _market_data_client
    if _market_data_client is None:
        _market_data_client = DeltaClient(base_url=config.MARKET_DATA_BASE_URL, api_key="", api_secret="")
    return _market_data_client


@app.route("/api/market-ticker")
@login_required
def api_market_ticker():
    """Live prices for the scrolling ticker strip. Public market data --
    same for every user, no personal API keys involved."""
    try:
        client = _get_market_data_client()
        tickers = client.get_tickers(COMMON_SYMBOLS)
        out = [
            {"symbol": t.get("symbol"), "price": t.get("close")}
            for t in tickers if t.get("close") is not None
        ]
        matched = {t["symbol"] for t in out}
        missing = [s for s in COMMON_SYMBOLS if s not in matched]
        if missing:
            app.logger.warning("Ticker symbols not found on Delta: %s", missing)
        return jsonify({"ok": True, "tickers": out, "fetched_at": int(time.time()), "missing": missing})
    except DeltaAPIError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Unexpected error: {e}"})


# ----------------------------------------------------------------------
# API: trade history + Excel export
# ----------------------------------------------------------------------
def _read_trade_history(username):
    path = os.path.join(users.user_dir(username), "trade_history.json")
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return []


@app.route("/api/trades")
@login_required
def api_trades():
    username = session["username"]
    history = _read_trade_history(username)
    # Most recent first, capped for the dashboard panel (export gets everything).
    recent = list(reversed(history))[:25]
    return jsonify({"ok": True, "trades": recent})


def _build_trade_history_workbook(rows_by_user):
    """
    rows_by_user: dict of username -> list of trade dicts (as recorded by
    the bot). Builds an .xlsx in memory with one sheet per user plus a
    combined "All Trades" summary sheet with a totals row using a real
    SUM formula (not a hardcoded number) so it recalculates if edited.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
    HEADER_FILL_COLOR = "1F2430"
    BODY_FONT = Font(name="Arial")
    CURRENCY_FMT = '$#,##0.00;($#,##0.00);-'

    columns = [
        ("Closed At", "closed_at"), ("Opened At", "opened_at"), ("Symbol", "symbol"),
        ("Direction", "direction"), ("Entry Price", "entry_price"), ("Exit Price", "exit_price"),
        ("Size", "size"), ("Stage at Close", "stage_at_close"), ("Close Reason", "close_reason"),
        ("Realized P&L (est.)", "realized_pnl_estimate"),
    ]

    def fmt_ts(ts):
        if ts is None:
            return None
        try:
            from datetime import datetime
            return datetime.utcfromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M:%S UTC")
        except (ValueError, TypeError):
            return None

    def write_sheet(ws, rows):
        for col_idx, (label, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor=HEADER_FILL_COLOR)
            cell.alignment = Alignment(horizontal="center")
        r = 2
        pnl_col = None
        for col_idx, (label, key) in enumerate(columns, start=1):
            if key == "realized_pnl_estimate":
                pnl_col = get_column_letter(col_idx)
        for row in rows:
            for col_idx, (label, key) in enumerate(columns, start=1):
                value = row.get(key)
                if key in ("closed_at", "opened_at"):
                    value = fmt_ts(value)
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.font = BODY_FONT
                if key == "realized_pnl_estimate" and value is not None:
                    cell.number_format = CURRENCY_FMT
            r += 1
        if rows and pnl_col:
            total_row = r
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
            total_cell = ws.cell(row=total_row, column=columns.index(("Realized P&L (est.)", "realized_pnl_estimate")) + 1)
            total_cell.value = f"=SUM({pnl_col}2:{pnl_col}{r - 1})"
            total_cell.font = Font(name="Arial", bold=True)
            total_cell.number_format = CURRENCY_FMT
        for col_idx, (label, _) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(label) + 2)

    wb = openpyxl.Workbook()
    all_ws = wb.active
    all_ws.title = "All Trades"
    all_rows = []
    for uname, rows in rows_by_user.items():
        for row in rows:
            all_rows.append(dict(row, username=uname))
    all_rows.sort(key=lambda r: r.get("closed_at") or 0)
    write_sheet(all_ws, all_rows)

    for uname, rows in rows_by_user.items():
        rows_sorted = sorted(rows, key=lambda r: r.get("closed_at") or 0)
        sheet_name = uname[:31] or "user"  # Excel sheet name length limit
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, rows_sorted)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/api/trades/export")
@login_required
def api_trades_export():
    """Exports the logged-in user's own trade history as an .xlsx."""
    username = session["username"]
    history = _read_trade_history(username)
    buf = _build_trade_history_workbook({username: history})
    return send_file(
        buf, as_attachment=True,
        download_name=f"aurelius-algo-trades-{username}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/trades/export-all")
@admin_required
def admin_trades_export_all():
    """Exports every user's trade history into one workbook, one sheet
    per user plus a combined summary sheet -- for oversight of everyone
    using the algo."""
    rows_by_user = {}
    for uname, _ in users.list_all_users():
        rows_by_user[uname] = _read_trade_history(uname)
    buf = _build_trade_history_workbook(rows_by_user)
    return send_file(
        buf, as_attachment=True,
        download_name="aurelius-algo-all-trades.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------------------------------------------------
# Admin: logs across all users (touch-friendly viewer)
# ----------------------------------------------------------------------
@app.route("/admin/logs")
@admin_required
def admin_logs_page():
    all_usernames = [u for u, _ in users.list_all_users()]
    return render_template(
        "admin_logs.html", username=session["username"], all_usernames=all_usernames,
    )


@app.route("/admin/api/logs/<target_username>")
@admin_required
def admin_api_logs(target_username):
    target_username = target_username.strip().lower()
    if not users.user_exists(target_username):
        return jsonify({"ok": False, "error": "No such user"}), 404
    paths = user_paths(target_username)
    since = request.args.get("since", default=0, type=int)
    log_path = paths["log"]
    if not os.path.exists(log_path):
        return jsonify({"ok": True, "lines": [], "offset": 0})
    size = os.path.getsize(log_path)
    if since > size:
        since = 0
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        f.seek(since)
        data = f.read()
        new_offset = f.tell()
    lines = data.splitlines() if data else []
    return jsonify({"ok": True, "lines": lines, "offset": new_offset})
@app.route("/api/logs")
@login_required
def api_logs():
    username = session["username"]
    log_path = user_paths(username)["log"]
    since = request.args.get("since", default=0, type=int)
    if not os.path.exists(log_path):
        return jsonify({"lines": [], "offset": 0})
    size = os.path.getsize(log_path)
    if since > size:
        since = 0
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        f.seek(since)
        data = f.read()
        new_offset = f.tell()
    lines = data.splitlines() if data else []
    return jsonify({"lines": lines, "offset": new_offset})


# ----------------------------------------------------------------------
# API: start / stop (per-user isolated subprocess)
# ----------------------------------------------------------------------
@app.route("/api/start", methods=["POST"])
@login_required
def api_start():
    username = session["username"]
    with _processes_lock:
        entry = _bot_processes.get(username)
        if entry is not None and entry["process"].poll() is None:
            return jsonify({"ok": False, "error": "Your bot is already running"}), 400

    payload = request.get_json(force=True, silent=True) or {}
    symbol = str(payload.get("symbol", "BTCUSD")).strip().upper()
    live_trading = bool(payload.get("live_trading", False))
    try:
        fixed_size = int(payload.get("fixed_size", 0))
        if fixed_size < 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Lot size must be a whole number, 0 or more"}), 400

    settings = users.get_settings(username)
    if not settings["api_key"] or not settings["api_secret"]:
        return jsonify({"ok": False, "error": "Add your Delta API key/secret in Settings first"}), 400

    users.update_settings(username, {
        "symbol": symbol, "fixed_size": fixed_size, "dry_run": not live_trading,
    })

    paths = user_paths(username)
    if os.path.exists(paths["stop_flag"]):
        os.remove(paths["stop_flag"])

    child_env = os.environ.copy()
    child_env.update({
        "DELTA_API_KEY": settings["api_key"],
        "DELTA_API_SECRET": settings["api_secret"],
        "USE_TESTNET": "true" if settings["use_testnet"] else "false",
        "SYMBOL": symbol,
        "FIXED_SIZE": str(fixed_size),
        "DRY_RUN": "false" if live_trading else "true",
        "LOG_FILE": paths["log_rel"],
        "STOP_FLAG_FILE": paths["stop_flag_rel"],
    })

    creationflags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
    try:
        process = subprocess.Popen(
            [sys.executable, BOT_SCRIPT], cwd=BASE_DIR, env=child_env, creationflags=creationflags,
        )
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not start bot: {e}"}), 500

    with _processes_lock:
        _bot_processes[username] = {"process": process, "stop_requested_at": None}

    return jsonify({"ok": True})


@app.route("/api/stop", methods=["POST"])
@login_required
def api_stop():
    username = session["username"]
    with _processes_lock:
        entry = _bot_processes.get(username)
        if entry is None or entry["process"].poll() is not None:
            _bot_processes.pop(username, None)
            return jsonify({"ok": True, "already_stopped": True})
        entry["stop_requested_at"] = time.time()

    paths = user_paths(username)
    with open(paths["stop_flag"], "w") as f:
        f.write("stop")
    return jsonify({"ok": True})


# ----------------------------------------------------------------------
# API: per-user settings (API keys + strategy account prefs)
# ----------------------------------------------------------------------
@app.route("/api/settings", methods=["GET", "POST"])
@login_required
def api_settings():
    username = session["username"]
    if request.method == "GET":
        settings = users.get_settings(username)
        return jsonify({
            "symbol": settings["symbol"],
            "fixed_size": settings["fixed_size"],
            "use_testnet": settings["use_testnet"],
            "api_key_masked": users.mask_secret(settings["api_key"]),
            "has_api_keys": bool(settings["api_key"] and settings["api_secret"]),
        })

    payload = request.get_json(force=True, silent=True) or {}
    updates = {}
    if "api_key" in payload and payload["api_key"].strip():
        updates["api_key"] = payload["api_key"].strip()
    if "api_secret" in payload and payload["api_secret"].strip():
        updates["api_secret"] = payload["api_secret"].strip()
    if "use_testnet" in payload:
        updates["use_testnet"] = bool(payload["use_testnet"])
    users.update_settings(username, updates)
    return jsonify({"ok": True})


# ----------------------------------------------------------------------
# Password changes
# ----------------------------------------------------------------------
@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    """Self-service: any logged-in user (including the admin) can change
    their own password, given the correct current one."""
    username = session["username"]
    payload = request.get_json(force=True, silent=True) or {}
    current_password = payload.get("current_password", "")
    new_password = payload.get("new_password", "")
    confirm_password = payload.get("confirm_password", "")

    if new_password != confirm_password:
        return jsonify({"ok": False, "error": "New passwords do not match"}), 400

    ok, error = users.change_password(username, current_password, new_password)
    if not ok:
        return jsonify({"ok": False, "error": error}), 400
    return jsonify({"ok": True})


@app.route("/admin/api/reset-password", methods=["POST"])
@admin_required
def admin_api_reset_password():
    """Admin override: reset any user's password without their current
    one, for someone who's locked out."""
    payload = request.get_json(force=True, silent=True) or {}
    target = payload.get("username", "").strip().lower()
    new_password = payload.get("new_password", "")
    confirm_password = payload.get("confirm_password", "")

    if new_password != confirm_password:
        return jsonify({"ok": False, "error": "New passwords do not match"}), 400

    ok, error = users.admin_reset_password(target, new_password)
    if not ok:
        return jsonify({"ok": False, "error": error}), 400
    return jsonify({"ok": True})


# ----------------------------------------------------------------------
# Contact support (users submit, only the admin can read the inbox)
# ----------------------------------------------------------------------
@app.route("/support", methods=["GET", "POST"])
@login_required
def support_page():
    username = session["username"]
    if request.method == "POST":
        message = request.form.get("message", "")
        ok, error = support.submit_message(username, message)
        my_messages = support.list_messages_for_user(username)
        return render_template(
            "support.html", username=username,
            is_admin=users.is_admin(username),
            my_messages=my_messages,
            sent=ok, error=None if ok else error,
        )
    my_messages = support.list_messages_for_user(username)
    return render_template(
        "support.html", username=username,
        is_admin=users.is_admin(username),
        my_messages=my_messages, sent=False, error=None,
    )


@app.route("/admin/support")
@admin_required
def admin_support_page():
    all_messages = support.list_all_messages()
    return render_template("admin_support.html", username=session["username"], messages=all_messages)


@app.route("/admin/api/support/resolve", methods=["POST"])
@admin_required
def admin_api_support_resolve():
    payload = request.get_json(force=True, silent=True) or {}
    message_id = payload.get("id")
    resolved = payload.get("resolved", True)
    try:
        message_id = int(message_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid message id"}), 400
    ok = support.set_resolved(message_id, resolved)
    return jsonify({"ok": ok})


# ----------------------------------------------------------------------
# Admin: live overview of every user (running state, mode, recent errors)
# ----------------------------------------------------------------------
@app.route("/admin/api/overview")
@admin_required
def admin_api_overview():
    overview = []
    unresolved_support = sum(1 for m in support.list_all_messages() if not m["resolved"])
    for uname, data in users.list_all_users():
        with _processes_lock:
            entry = _bot_processes.get(uname)
            running = entry is not None and entry["process"].poll() is None
        settings = users.get_settings(uname)
        paths = user_paths(uname)
        recent_errors = 0
        last_error_line = None
        if os.path.exists(paths["log"]):
            try:
                with open(paths["log"], "r", encoding="utf-8", errors="replace") as f:
                    # Only scan the tail of the file -- avoids reading a
                    # potentially large log in full on every poll.
                    f.seek(max(0, os.path.getsize(paths["log"]) - 60000))
                    tail_lines = f.read().splitlines()
                for line in tail_lines:
                    if "[ERROR]" in line:
                        recent_errors += 1
                        last_error_line = line.strip()
            except OSError:
                pass
        overview.append({
            "username": uname,
            "is_admin": data.get("is_admin", False),
            "approved": data.get("approved", False),
            "suspended": data.get("suspended", False),
            "running": running,
            "symbol": settings["symbol"],
            "dry_run": settings["dry_run"],
            "use_testnet": settings["use_testnet"],
            "recent_errors": recent_errors,
            "last_error": last_error_line,
        })
    return jsonify({"ok": True, "users": overview, "unresolved_support": unresolved_support})


def _background_watchdog():
    while True:
        with _processes_lock:
            for username, entry in list(_bot_processes.items()):
                if (entry["stop_requested_at"] is not None
                        and entry["process"].poll() is None
                        and (time.time() - entry["stop_requested_at"]) > 15):
                    entry["process"].terminate()
        time.sleep(2)


if __name__ == "__main__":
    with open(os.path.join(BASE_DIR, "server.pid"), "w") as f:
        f.write(str(os.getpid()))
    threading.Thread(target=_background_watchdog, daemon=True).start()
    app.run(host="0.0.0.0", port=5000, debug=False)
